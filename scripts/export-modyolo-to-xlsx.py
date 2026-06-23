"""
Export modyolo-export-merged.csv to .xlsx format.
Each record in a row, each field in a column.
Adds downloadUrl/1 column (URL + /1) like liteapks_formatted_data.xlsx.
"""
import csv
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
CSV_PATH = SCRIPT_DIR / "modyolo-export-merged.csv"
XLSX_PATH = SCRIPT_DIR / "modyolo_formatted_data.xlsx"

COLUMNS = [
    "slug",
    "title",
    "type",
    "description",
    "appName",
    "publisher",
    "genre",
    "size",
    "version",
    "modInfo",
    "modFeatures",
    "downloadNotes",
    "downloadUrl/1",
    "downloadVersions",
    "bannerImage",
    "iconImage",
    "categorySlug",
    "contentBlocks",
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MODYOLO"

    # Header row
    for col, name in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col, value=name)

    with open(CSV_PATH, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row_idx, row in enumerate(reader, start=2):
            download_url = (row.get("downloadUrl") or "").strip()
            if download_url and not download_url.endswith("/1"):
                download_url = download_url.rstrip("/") + "/1"

            values = [
                row.get("slug", ""),
                row.get("title", ""),
                row.get("type", ""),
                row.get("description", ""),
                row.get("appName", ""),
                row.get("publisher", ""),
                row.get("genre", ""),
                row.get("size", ""),
                row.get("version", ""),
                row.get("modInfo", ""),
                row.get("modFeatures", ""),
                row.get("downloadNotes", ""),
                download_url,
                row.get("downloadVersions", ""),
                row.get("bannerImage", ""),
                row.get("iconImage", ""),
                row.get("categorySlug", ""),
                row.get("contentBlocks", ""),
            ]
            for col, val in enumerate(values, 1):
                ws.cell(row=row_idx, column=col, value=val or "")

    wb.save(XLSX_PATH)
    print(f"Exported {row_idx - 1} records to {XLSX_PATH}")


if __name__ == "__main__":
    main()
