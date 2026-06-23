"""Add /1 at the end of URL in column M of liteapks_formatted_data.xlsx"""
import openpyxl
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = SCRIPT_DIR / "liteapks_formatted_data.xlsx"

wb = openpyxl.load_workbook(XLSX_PATH)
for sheet in wb.worksheets:
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=13)  # column M = 13
        val = cell.value
        if val is not None and isinstance(val, str) and val.strip():
            val = val.strip()
            if not val.endswith("/1"):
                cell.value = val.rstrip("/") + "/1"
wb.save(XLSX_PATH)
print(f"Updated column M in {XLSX_PATH}")
