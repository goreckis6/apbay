"""
Download APK files from URLs in column M of liteapks_formatted_data.xlsx.

For each URL, opens the page, finds the download link (#download-loaded-link),
and downloads the APK file using the exact filename from the download URL.
Writes a manifest CSV (filename, slug, title) for easy matching.

Usage:
  py scripts/download-apks-from-excel.py [--output-dir downloads] [--resume] [--limit N]

Requires: openpyxl, playwright
  py -m pip install openpyxl playwright
  py -m playwright install chromium
"""

import argparse
import csv
import re
import time
from pathlib import Path
from urllib.parse import urlparse, unquote

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = SCRIPT_DIR / "liteapks_formatted_data.xlsx"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "downloaded_apks"
MANIFEST_NAME = "download-manifest.csv"

# Excel columns: A=slug, B=title, M=downloadUrl
COL_SLUG = 1
COL_TITLE = 2
COL_URL = 13


def get_rows_from_excel(xlsx_path: Path) -> list[tuple[int, str, str, str]]:
    """Return list of (row_index, slug, title, url) for rows with URLs in column M."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets:
        for row in range(1, sheet.max_row + 1):
            url_cell = sheet.cell(row=row, column=COL_URL)
            url_val = url_cell.value
            if url_val is not None and isinstance(url_val, str) and url_val.strip():
                url = url_val.strip()
                if url.startswith("http"):
                    slug = sheet.cell(row=row, column=COL_SLUG).value
                    title = sheet.cell(row=row, column=COL_TITLE).value
                    slug = str(slug).strip() if slug else ""
                    title = str(title).strip() if title else ""
                    rows.append((row, slug, title, url))
    wb.close()
    return rows


def exact_filename_from_url(url: str, fallback: str = "apk") -> str:
    """Extract exact filename from download URL (preserves original name, sanitizes for FS)."""
    parsed = urlparse(url)
    path = unquote(parsed.path)
    m = re.search(r"/([^/]+\.apk)(?:\?|$)", path, re.I)
    if m:
        name = m.group(1)
    else:
        name = path.rstrip("/").split("/")[-1] or fallback
        if not name.lower().endswith(".apk"):
            name = f"{name}.apk"
    name = re.sub(r'[<>:"/\\|?*]', "_", name)
    return name[:200] if len(name) > 200 else name


def download_file_via_requests(url: str, save_path: Path) -> bool:
    """Download file from URL using requests."""
    try:
        import urllib.request
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        })
        with urllib.request.urlopen(req, timeout=120) as resp:
            save_path.write_bytes(resp.read())
        return True
    except Exception as e:
        print(f"  Download error: {e}")
        return False


def get_download_href_from_page(page_url: str, browser_context) -> str | None:
    """
    Open page_url, find #download-loaded-link, return its href.
    """
    page = browser_context.new_page()
    page.set_default_timeout(30000)
    try:
        page.goto(page_url, wait_until="networkidle")
        # Primary: #download-loaded-link; fallback: a.download or a.btn-primary.download
        link = page.locator("#download-loaded-link, a.download[href], a.btn-primary.download[href]").first
        link.wait_for(state="visible", timeout=15000)
        href = link.get_attribute("href")
        if not href or not href.startswith("http"):
            data_href = link.get_attribute("data-href")
            if data_href:
                import base64
                try:
                    href = base64.b64decode(data_href).decode("utf-8")
                except Exception:
                    pass
        return href if (href and href.startswith("http")) else None
    except PlaywrightTimeout as e:
        print(f"  Timeout: {e}")
        return None
    except Exception as e:
        print(f"  Error: {e}")
        return None
    finally:
        page.close()


def download_apk_from_page(
    page_url: str,
    output_dir: Path,
    browser_context,
    slug: str,
    title: str,
    resume: bool = False,
) -> tuple[str, Path, bool] | None:
    """
    Open page_url, find #download-loaded-link, get href, download APK with exact filename.
    Returns (filename, save_path, was_skipped) on success, None on failure.
    """
    href = get_download_href_from_page(page_url, browser_context)
    if not href:
        return None

    filename = exact_filename_from_url(href, f"apk_{int(time.time())}")
    save_path = output_dir / filename
    if resume and save_path.exists():
        print(f"  Skipped (exists): {filename}")
        return (filename, save_path, True)
    if download_file_via_requests(href, save_path):
        return (filename, save_path, False)
    return None


def main():
    parser = argparse.ArgumentParser(description="Download APKs from column M URLs in Excel")
    parser.add_argument("--output-dir", "-o", default=str(DEFAULT_OUTPUT_DIR), help="Output directory for APKs")
    parser.add_argument("--resume", action="store_true", help="Skip URLs whose files already exist")
    parser.add_argument("--limit", "-n", type=int, default=0, help="Max number of downloads (0 = all)")
    parser.add_argument("--delay", type=float, default=2.0, help="Delay between requests (seconds)")
    args = parser.parse_args()

    if not XLSX_PATH.exists():
        print(f"Error: {XLSX_PATH} not found")
        return 1

    rows = get_rows_from_excel(XLSX_PATH)
    print(f"Found {len(rows)} rows with URLs in column M")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    manifest_path = output_dir / MANIFEST_NAME
    manifest_exists = manifest_path.exists()
    manifest_file = open(manifest_path, "a", newline="", encoding="utf-8")
    manifest_writer = csv.writer(manifest_file)
    if not manifest_exists:
        manifest_writer.writerow(["filename", "slug", "title", "page_url"])

    done_count = 0
    fail_count = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )
        context.set_default_timeout(30000)

        for i, (row, slug, title, url) in enumerate(rows):
            if args.limit and done_count >= args.limit:
                print(f"Reached limit of {args.limit}")
                break

            print(f"[{i + 1}/{len(rows)}] Row {row} ({slug}): {url[:50]}...")
            result = download_apk_from_page(url, output_dir, context, slug, title, resume=args.resume)
            if result:
                filename, save_path, was_skipped = result
                if not was_skipped:
                    print(f"  Saved: {filename}")
                    manifest_writer.writerow([filename, slug, title, url])
                    manifest_file.flush()
                done_count += 1
            else:
                fail_count += 1

            if args.delay and i < len(rows) - 1:
                time.sleep(args.delay)

        browser.close()

    manifest_file.close()
    print(f"\nDone. Downloaded: {done_count}, Failed: {fail_count}")
    print(f"Manifest: {manifest_path}")
    return 0


if __name__ == "__main__":
    exit(main())
